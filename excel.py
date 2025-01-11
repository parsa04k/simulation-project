import openpyxl
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import json  # We'll use this to convert lists of dictionaries to a JSON-like string

file_name = r'C:\Users\ASUS\Music\ترم پنج\شبیه سازی\code\output.xlsx'

def xlsx_writer(data_dict, sheet_name, file_name=file_name):
    # Check if the file exists and handle errors
    if not os.path.exists(file_name):
        workbook = openpyxl.Workbook()  # Create a new workbook
    else:
        try:
            workbook = openpyxl.load_workbook(file_name)
        except Exception as e:
            print(f"Error loading the workbook: {e}")
            return  # Exit if there's an error loading the file
    
    # Add a new sheet or select the existing sheet
    if sheet_name not in workbook.sheetnames:
        sheet = workbook.create_sheet(sheet_name)
    else:
        sheet = workbook[sheet_name]
    
    # Extract columns and values from the dictionary
    columns = list(data_dict.keys())
    values = list(data_dict.values())
    
    # Add column headers
    for col_num, column in enumerate(columns, 1):
        sheet.cell(row=1, column=col_num, value=column)
    
    # Add values under each column, including lists of dictionaries
    for row_num, row in enumerate(zip(*values), start=2):
        for col_num, value in enumerate(row, 1):
            # Check if the value is a list of dictionaries
            if isinstance(value, list) and isinstance(value[0], dict):
                # Convert the list of dictionaries to a string (JSON-like)
                value = json.dumps(value)  # Convert list to a JSON string representation
            sheet.cell(row=row_num, column=col_num, value=value)
    
    # Align all cells to the center
    for row in sheet.iter_rows(min_row=1, min_col=1, max_col=len(columns), max_row=sheet.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Apply borders to all cells in the sheet
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))

    for row in sheet.iter_rows(min_row=1, min_col=1, max_col=len(columns), max_row=sheet.max_row):
        for cell in row:
            cell.border = thin_border
    
    for row in sheet.iter_rows(min_row=1, min_col=1, max_col=len(columns), max_row=len(values) + 1):
        for cell in row:
            cell.border = thin_border
    
    # Auto width the columns, except the last one
    for col_num in range(1, len(columns)):
        max_length = 0
        column = get_column_letter(col_num)
        for row in sheet.iter_rows(min_row=1, max_row=len(values) + 1, min_col=col_num, max_col=col_num):
            for cell in row:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width
    
    # Freeze the first row
    sheet.freeze_panes = "A2"
    
    # Save the workbook
    workbook.save(file_name)

import scipy.stats as stats

def mean_and_confidence_interval(data, confidence=0.95):
    # Calculate the mean and standard error of the mean
    mean = sum(data)/len(data)
    std = (sum((i-mean)**2 for i in data)/(len(data)-1))**0.5
    # Calculate the margin of error
    margin_of_error = (std/len(data)) * stats.t.ppf((1 + confidence) / 2, len(data) - 1)
    
    # Calculate the confidence interval
    lower_bound = mean - margin_of_error
    upper_bound = mean + margin_of_error
    
    return mean, (lower_bound, upper_bound)

